# apy/view/informes/views.py — MÓDULO INFORMES COMPLETO

from django.utils import timezone
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache
from django.utils.decorators import method_decorator
from django.views import View
from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncMonth, TruncWeek
from django.urls import reverse_lazy
from django.http import HttpResponse
from datetime import date, timedelta, datetime
import json

from apy.models import (
    DetalleServicio, Pagos, Nomina, Profile,
    Module, Permission
)
from apy.decorators import PermisoRequeridoMixin


# ─────────────────────────────────────────────────────────────
# HELPERS DE FECHA
# ─────────────────────────────────────────────────────────────
def get_rango_fechas(request):
    """
    Lee los parámetros GET y devuelve (fecha_inicio, fecha_fin, periodo_activo).
    Prioridad: rango personalizado > periodo rápido > mes actual por defecto.
    """
    hoy = timezone.now().date()

    # Rango personalizado
    desde_str = request.GET.get('desde', '')
    hasta_str = request.GET.get('hasta', '')
    if desde_str and hasta_str:
        try:
            desde = date.fromisoformat(desde_str)
            hasta = date.fromisoformat(hasta_str)
            return desde, hasta, 'personalizado'
        except ValueError:
            pass

    # Periodo rápido
    periodo = request.GET.get('periodo', 'mes')
    if periodo == 'semana':
        desde = hoy - timedelta(days=hoy.weekday())   # lunes de esta semana
        hasta = desde + timedelta(days=6)
        return desde, hasta, 'semana'
    elif periodo == 'anio':
        desde = date(hoy.year, 1, 1)
        hasta = date(hoy.year, 12, 31)
        return desde, hasta, 'anio'
    else:  # mes (default)
        desde = date(hoy.year, hoy.month, 1)
        # último día del mes
        if hoy.month == 12:
            hasta = date(hoy.year, 12, 31)
        else:
            hasta = date(hoy.year, hoy.month + 1, 1) - timedelta(days=1)
        return desde, hasta, 'mes'


def labels_para_grafica(desde, hasta, qs_fecha_field):
    """Genera labels de días para gráficas pequeñas (máx 31 días)."""
    delta = (hasta - desde).days
    if delta <= 31:
        return [(desde + timedelta(days=i)).strftime('%d/%m') for i in range(delta + 1)]
    return None


# ─────────────────────────────────────────────────────────────
# VISTA PRINCIPAL DEL DASHBOARD DE INFORMES
# ─────────────────────────────────────────────────────────────
@never_cache
@login_required(login_url=reverse_lazy('login:login'))
def InformeDashboardView(request):
    desde, hasta, periodo_activo = get_rango_fechas(request)
    estado_filtro = request.GET.get('estado', 'todos')

    # ── SERVICIOS ─────────────────────────────────────────────
    qs_servicios = DetalleServicio.objects.filter(
        estado=True,
        fecha_creacion__date__gte=desde,
        fecha_creacion__date__lte=hasta,
    )
    if estado_filtro != 'todos':
        qs_servicios = qs_servicios.filter(proceso=estado_filtro)

    total_servicios     = qs_servicios.count()
    servicios_terminados = qs_servicios.filter(proceso='terminado').count()
    servicios_en_proceso = qs_servicios.filter(proceso='proceso').count()
    ingresos_servicios  = sum(float(s.subtotal) for s in qs_servicios.filter(proceso='terminado'))

    # Datos para gráfica servicios por día
    servicios_por_dia = []
    labels_servicios  = []
    delta = (hasta - desde).days
    if delta <= 62:
        for i in range(delta + 1):
            dia = desde + timedelta(days=i)
            cnt = qs_servicios.filter(fecha_creacion__date=dia).count()
            servicios_por_dia.append(cnt)
            labels_servicios.append(dia.strftime('%d/%m'))
    else:
        # Agrupado por mes si el rango es mayor a 2 meses
        meses_data = (
            qs_servicios
            .annotate(mes=TruncMonth('fecha_creacion'))
            .values('mes')
            .annotate(total=Count('id'))
            .order_by('mes')
        )
        for m in meses_data:
            labels_servicios.append(m['mes'].strftime('%b %Y'))
            servicios_por_dia.append(m['total'])

    # ── PAGOS A PROVEEDORES ───────────────────────────────────
    qs_pagos = Pagos.objects.filter(
        estado=True,
        fecha__gte=desde,
        fecha__lte=hasta,
    )
    total_pagos   = qs_pagos.count()
    monto_pagos   = qs_pagos.aggregate(t=Sum('monto_total'))['t'] or 0

    # Datos para gráfica pagos por día
    pagos_por_dia  = []
    labels_pagos   = []
    if delta <= 62:
        for i in range(delta + 1):
            dia = desde + timedelta(days=i)
            monto = qs_pagos.filter(fecha=dia).aggregate(t=Sum('monto_total'))['t'] or 0
            pagos_por_dia.append(float(monto))
            labels_pagos.append(dia.strftime('%d/%m'))
    else:
        meses_pagos = (
            qs_pagos
            .annotate(mes=TruncMonth('fecha'))
            .values('mes')
            .annotate(total=Sum('monto_total'))
            .order_by('mes')
        )
        for m in meses_pagos:
            labels_pagos.append(m['mes'].strftime('%b %Y'))
            pagos_por_dia.append(float(m['total'] or 0))

    # ── NÓMINA ────────────────────────────────────────────────
    qs_nomina = Nomina.objects.filter(
        fecha_pago__gte=desde,
        fecha_pago__lte=hasta,
    )
    total_nomina_registros = qs_nomina.count()
    monto_nomina           = qs_nomina.aggregate(t=Sum('monto'))['t'] or 0

    # Datos para gráfica nómina por día
    nomina_por_dia = []
    labels_nomina  = []
    if delta <= 62:
        for i in range(delta + 1):
            dia = desde + timedelta(days=i)
            monto = qs_nomina.filter(fecha_pago=dia).aggregate(t=Sum('monto'))['t'] or 0
            nomina_por_dia.append(float(monto))
            labels_nomina.append(dia.strftime('%d/%m'))
    else:
        meses_nomina = (
            qs_nomina
            .annotate(mes=TruncMonth('fecha_pago'))
            .values('mes')
            .annotate(total=Sum('monto'))
            .order_by('mes')
        )
        for m in meses_nomina:
            labels_nomina.append(m['mes'].strftime('%b %Y'))
            nomina_por_dia.append(float(m['total'] or 0))

    # ── TABLAS DE DETALLE ─────────────────────────────────────
    servicios_tabla = qs_servicios.select_related(
        'id_vehiculo', 'empleado', 'cliente'
    ).order_by('-fecha_creacion')[:20]

    pagos_tabla = qs_pagos.select_related('proveedor').order_by('-fecha')[:20]

    nomina_tabla = qs_nomina.select_related(
        'empleado__user'
    ).order_by('-fecha_pago')[:20]

    context = {
        # Rango activo
        'periodo_activo': periodo_activo,
        'desde': desde.strftime('%Y-%m-%d'),
        'hasta': hasta.strftime('%Y-%m-%d'),
        'desde_display': desde.strftime('%d/%m/%Y'),
        'hasta_display': hasta.strftime('%d/%m/%Y'),
        'estado_filtro': estado_filtro,

        # Métricas servicios
        'total_servicios':      total_servicios,
        'servicios_terminados': servicios_terminados,
        'servicios_en_proceso': servicios_en_proceso,
        'ingresos_servicios':   round(ingresos_servicios, 0),

        # Métricas pagos
        'total_pagos':  total_pagos,
        'monto_pagos':  float(monto_pagos),

        # Métricas nómina
        'total_nomina_registros': total_nomina_registros,
        'monto_nomina':           float(monto_nomina),

        # Gráficas (JSON safe)
        'labels_servicios':  json.dumps(labels_servicios),
        'datos_servicios':   json.dumps(servicios_por_dia),
        'labels_pagos':      json.dumps(labels_pagos),
        'datos_pagos':       json.dumps(pagos_por_dia),
        'labels_nomina':     json.dumps(labels_nomina),
        'datos_nomina':      json.dumps(nomina_por_dia),

        # Tablas
        'servicios_tabla': servicios_tabla,
        'pagos_tabla':     pagos_tabla,
        'nomina_tabla':    nomina_tabla,
    }
    return render(request, 'informes/dashboard_informes.html', context)


# ─────────────────────────────────────────────────────────────
# VISTAS EXISTENTES (mantener compatibilidad con urls.py)
# ─────────────────────────────────────────────────────────────
class InformeListView(PermisoRequeridoMixin, View):
    module_name = 'Informes'
    permission_required = 'view'

    def get(self, request):
        from apy.models import Informes
        informes = Informes.objects.filter(estado=True).select_related(
            'detalle_servicio__id_vehiculo', 'id_empleado__user'
        ).order_by('-fecha')
        return render(request, 'informes/lista_informes.html', {'informes': informes})


class CreateInformeView(PermisoRequeridoMixin, View):
    module_name = 'Informes'
    permission_required = 'add'

    def get(self, request):
        from apy.forms import InformeForm
        form = InformeForm()
        servicios = DetalleServicio.objects.filter(
            estado=True, proceso='terminado'
        ).exclude(informes__isnull=False)
        return render(request, 'informes/crear_informe.html', {
            'form': form,
            'servicios': servicios,
        })

    def post(self, request):
        from apy.forms import InformeForm
        from apy.models import Informes
        form = InformeForm(request.POST)
        if form.is_valid():
            informe = form.save(commit=False)
            servicio_id = request.POST.get('detalle_servicio')
            try:
                informe.detalle_servicio = DetalleServicio.objects.get(pk=servicio_id)
                informe.id_empleado = request.user.profile
                informe.save()
                from django.contrib import messages
                messages.success(request, 'Informe creado correctamente.')
                return __import__('django.shortcuts', fromlist=['redirect']).redirect('apy:informe_lista')
            except Exception as e:
                form.add_error(None, str(e))
        servicios = DetalleServicio.objects.filter(estado=True, proceso='terminado')
        return render(request, 'informes/crear_informe.html', {'form': form, 'servicios': servicios})


class InformeDetailView(PermisoRequeridoMixin, View):
    module_name = 'Informes'
    permission_required = 'view'

    def get(self, request, pk):
        from apy.models import Informes
        from django.shortcuts import get_object_or_404
        informe = get_object_or_404(Informes, pk=pk)
        return render(request, 'informes/detalle_informe.html', {'informe': informe})


class InformeGerencialView(PermisoRequeridoMixin, View):
    module_name = 'Informes'
    permission_required = 'view'

    def get(self, request):
        return render(request, 'informes/informe_gerencial.html', {})


def exportar_informe_excel(request):
    """Exporta el informe filtrado a Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    desde, hasta, periodo = get_rango_fechas(request)
    estado_filtro = request.GET.get('estado', 'todos')
    modulo = request.GET.get('modulo', 'servicios')

    wb = openpyxl.Workbook()
    ws = wb.active

    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    if modulo == 'servicios':
        ws.title = 'Servicios'
        qs = DetalleServicio.objects.filter(
            estado=True,
            fecha_creacion__date__gte=desde,
            fecha_creacion__date__lte=hasta,
        )
        if estado_filtro != 'todos':
            qs = qs.filter(proceso=estado_filtro)

        headers = ['ID', 'Vehículo', 'Cliente', 'Empleado', 'Fecha', 'Estado', 'Total']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for s in qs.select_related('id_vehiculo', 'cliente', 'empleado').order_by('-fecha_creacion'):
            ws.append([
                s.id,
                str(s.id_vehiculo.placa),
                str(s.cliente.nombre),
                str(s.empleado.user.get_full_name() if s.empleado else '—'),
                s.fecha_creacion.strftime('%d/%m/%Y'),
                s.get_proceso_display(),
                float(s.subtotal),
            ])

    elif modulo == 'pagos':
        ws.title = 'Pagos'
        qs = Pagos.objects.filter(estado=True, fecha__gte=desde, fecha__lte=hasta)
        headers = ['ID', 'Proveedor', 'Fecha', 'Tipo Pago', 'Monto Total']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for p in qs.select_related('proveedor').order_by('-fecha'):
            ws.append([p.id_pago, p.proveedor.nombre, p.fecha.strftime('%d/%m/%Y'), p.tipo_pago, float(p.monto_total)])

    elif modulo == 'nomina':
        ws.title = 'Nómina'
        qs = Nomina.objects.filter(fecha_pago__gte=desde, fecha_pago__lte=hasta)
        headers = ['ID', 'Empleado', 'Fecha Pago', 'Monto']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for n in qs.select_related('empleado__user').order_by('-fecha_pago'):
            ws.append([
                n.id,
                n.empleado.user.get_full_name() or n.empleado.user.username,
                n.fecha_pago.strftime('%d/%m/%Y'),
                float(n.monto),
            ])

    # Ajustar anchos
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="informe_{modulo}_{desde}_{hasta}.xlsx"'
    wb.save(response)
    return response