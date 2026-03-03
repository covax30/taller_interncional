# apy/view/Estadisticas/views.py

from datetime import timedelta, date
import json

from django.utils import timezone
from django.shortcuts import render
from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncMonth
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache
from django.urls import reverse_lazy
from django.http import JsonResponse, HttpResponse

from apy.models import (
    Caja, Cliente, DetalleServicio, Gastos, Insumos,
    Nomina, Pagos, Vehiculo,
    DetalleRepuesto, DetalleInsumos,
)


# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────
def es_gerente(user):
    return user.is_superuser or user.is_staff


def get_rango_fechas(request):
    hoy = timezone.now().date()

    desde_str = request.GET.get('desde', '')
    hasta_str = request.GET.get('hasta', '')
    if desde_str and hasta_str:
        try:
            desde = date.fromisoformat(desde_str)
            hasta = date.fromisoformat(hasta_str)
            return desde, hasta, 'personalizado'
        except ValueError:
            pass

    periodo = request.GET.get('periodo', 'mes')
    if periodo == 'semana':
        desde = hoy - timedelta(days=hoy.weekday())
        hasta = desde + timedelta(days=6)
        return desde, hasta, 'semana'
    elif periodo == 'anio':
        desde = date(hoy.year, 1, 1)
        hasta = date(hoy.year, 12, 31)
        return desde, hasta, 'anio'
    else:
        desde = date(hoy.year, hoy.month, 1)
        if hoy.month == 12:
            hasta = date(hoy.year, 12, 31)
        else:
            hasta = date(hoy.year, hoy.month + 1, 1) - timedelta(days=1)
        return desde, hasta, 'mes'


# ─────────────────────────────────────────────────────────────
# VISTA PRINCIPAL
# ─────────────────────────────────────────────────────────────
@never_cache
@login_required(login_url=reverse_lazy('login:login'))
def estadisticas(request):
    hoy         = timezone.now()
    mes_actual  = hoy.month
    anio_actual = hoy.year
    NOMBRES_MESES = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']

    # ── BLOQUE 1: Contadores principales ─────────────────────
    total_insumos   = Insumos.objects.filter(estado=True).count()
    total_vehiculos = Vehiculo.objects.filter(estado=True).count()
    total_clientes  = Cliente.objects.filter(estado=True).count()
    total_gastos    = Gastos.objects.filter(estado=True).count()

    # ── BLOQUE 2: Métricas de servicios del mes ───────────────
    servicios_mes = DetalleServicio.objects.filter(
        fecha_creacion__month=mes_actual,
        fecha_creacion__year=anio_actual,
        proceso='terminado',
        estado=True,
    )
    ingresos_mes             = sum(float(s.subtotal) for s in servicios_mes)
    servicios_terminados_mes = servicios_mes.count()
    servicios_en_proceso     = DetalleServicio.objects.filter(
        proceso='proceso', estado=True
    ).count()
    ticket_promedio = (
        round(ingresos_mes / servicios_terminados_mes, 0)
        if servicios_terminados_mes > 0 else 0
    )

    # Variación vs mes anterior
    mes_ant  = mes_actual - 1 if mes_actual > 1 else 12
    anio_ant = anio_actual if mes_actual > 1 else anio_actual - 1
    ingresos_mes_anterior = sum(
        float(s.subtotal)
        for s in DetalleServicio.objects.filter(
            fecha_creacion__month=mes_ant,
            fecha_creacion__year=anio_ant,
            proceso='terminado',
            estado=True,
        )
    )
    variacion_ingresos = (
        round(((ingresos_mes - ingresos_mes_anterior) / ingresos_mes_anterior) * 100, 1)
        if ingresos_mes_anterior > 0 else 0
    )

    # ── BLOQUE 3: Tarjetas financieras del mes ────────────────
    gastos_mes = float(
        Caja.objects.filter(
            fecha__month=mes_actual,
            fecha__year=anio_actual,
        ).exclude(tipo_movimiento='Ingreso')
        .aggregate(total=Sum('monto'))['total'] or 0
    )
    utilidad_mes = ingresos_mes - gastos_mes

    # ── BLOQUE 4: Gráfico ingresos vs gastos (6 meses) ───────
    meses_labels   = []
    datos_ingresos = []
    datos_gastos   = []
    for i in range(5, -1, -1):
        m = mes_actual - i
        a = anio_actual
        if m <= 0:
            m += 12
            a -= 1
        meses_labels.append(f"{NOMBRES_MESES[m-1]} {a}")

        ingreso_m = sum(
            float(s.subtotal)
            for s in DetalleServicio.objects.filter(
                fecha_creacion__month=m,
                fecha_creacion__year=a,
                proceso='terminado',
                estado=True,
            )
        )
        gasto_m = float(
            Caja.objects.filter(fecha__month=m, fecha__year=a)
            .exclude(tipo_movimiento='Ingreso')
            .aggregate(t=Sum('monto'))['t'] or 0
        )
        datos_ingresos.append(round(ingreso_m, 2))
        datos_gastos.append(round(gasto_m, 2))

    # ── BLOQUE 5: Servicios recientes ─────────────────────────
    servicios_recientes = (
        DetalleServicio.objects
        .filter(estado=True)
        .select_related('id_vehiculo', 'empleado', 'cliente')
        .order_by('-fecha_creacion')[:10]
    )

    # ── BLOQUE 6: Informes por período (filtros) ──────────────
    desde, hasta, periodo_activo = get_rango_fechas(request)
    estado_filtro = request.GET.get('estado', 'todos')
    delta = (hasta - desde).days

    qs_servicios = DetalleServicio.objects.filter(
        estado=True,
        fecha_creacion__date__gte=desde,
        fecha_creacion__date__lte=hasta,
    )
    if estado_filtro != 'todos':
        qs_servicios = qs_servicios.filter(proceso=estado_filtro)

    total_servicios_informe      = qs_servicios.count()
    servicios_terminados_informe = qs_servicios.filter(proceso='terminado').count()
    servicios_en_proceso_informe = qs_servicios.filter(proceso='proceso').count()
    ingresos_servicios_informe   = sum(
        float(s.subtotal) for s in qs_servicios.filter(proceso='terminado')
    )

    # Gráfica servicios del período
    labels_servicios  = []
    servicios_por_dia = []
    if delta <= 62:
        for i in range(delta + 1):
            dia = desde + timedelta(days=i)
            labels_servicios.append(dia.strftime('%d/%m'))
            servicios_por_dia.append(qs_servicios.filter(fecha_creacion__date=dia).count())
    else:
        for m in (
            qs_servicios
            .annotate(mes=TruncMonth('fecha_creacion'))
            .values('mes').annotate(total=Count('id')).order_by('mes')
        ):
            labels_servicios.append(m['mes'].strftime('%b %Y'))
            servicios_por_dia.append(m['total'])

    # Pagos del período
    qs_pagos    = Pagos.objects.filter(estado=True, fecha__gte=desde, fecha__lte=hasta)
    total_pagos = qs_pagos.count()
    monto_pagos = float(qs_pagos.aggregate(t=Sum('monto_total'))['t'] or 0)

    labels_pagos  = []
    pagos_por_dia = []
    if delta <= 62:
        for i in range(delta + 1):
            dia = desde + timedelta(days=i)
            labels_pagos.append(dia.strftime('%d/%m'))
            pagos_por_dia.append(
                float(qs_pagos.filter(fecha=dia).aggregate(t=Sum('monto_total'))['t'] or 0)
            )
    else:
        for m in (
            qs_pagos.annotate(mes=TruncMonth('fecha'))
            .values('mes').annotate(total=Sum('monto_total')).order_by('mes')
        ):
            labels_pagos.append(m['mes'].strftime('%b %Y'))
            pagos_por_dia.append(float(m['total'] or 0))

    # Nómina del período
    qs_nomina              = Nomina.objects.filter(fecha_pago__gte=desde, fecha_pago__lte=hasta)
    total_nomina_registros = qs_nomina.count()
    monto_nomina           = float(qs_nomina.aggregate(t=Sum('monto'))['t'] or 0)

    labels_nomina  = []
    nomina_por_dia = []
    if delta <= 62:
        for i in range(delta + 1):
            dia = desde + timedelta(days=i)
            labels_nomina.append(dia.strftime('%d/%m'))
            nomina_por_dia.append(
                float(qs_nomina.filter(fecha_pago=dia).aggregate(t=Sum('monto'))['t'] or 0)
            )
    else:
        for m in (
            qs_nomina.annotate(mes=TruncMonth('fecha_pago'))
            .values('mes').annotate(total=Sum('monto')).order_by('mes')
        ):
            labels_nomina.append(m['mes'].strftime('%b %Y'))
            nomina_por_dia.append(float(m['total'] or 0))

    # Tablas del período
    servicios_tabla = (
        qs_servicios
        .select_related('id_vehiculo', 'empleado', 'cliente')
        .order_by('-fecha_creacion')[:20]
    )
    pagos_tabla  = qs_pagos.select_related('proveedor').order_by('-fecha')[:20]
    nomina_tabla = qs_nomina.select_related('empleado__user').order_by('-fecha_pago')[:20]

    # ── CONTEXTO ──────────────────────────────────────────────
    context = {
        # Contadores
        'total_insumos':   total_insumos,
        'total_vehiculos': total_vehiculos,
        'total_clientes':  total_clientes,
        'total_gastos':    total_gastos,

        # Métricas servicios
        'servicios_en_proceso':     servicios_en_proceso,
        'servicios_terminados_mes': servicios_terminados_mes,
        'ticket_promedio':          round(ticket_promedio, 0),
        'variacion_ingresos':       variacion_ingresos,

        # Financiero del mes
        'ingresos_mes': round(ingresos_mes, 0),
        'gastos_mes':   round(gastos_mes, 0),
        'utilidad_mes': round(utilidad_mes, 0),

        # Gráfico 6 meses
        'meses':          json.dumps(meses_labels),
        'datos_ingresos': json.dumps(datos_ingresos),
        'datos_gastos':   json.dumps(datos_gastos),

        # Servicios recientes
        'servicios_recientes': servicios_recientes,

        # Informes / filtros
        'periodo_activo': periodo_activo,
        'desde':          desde.strftime('%Y-%m-%d'),
        'hasta':          hasta.strftime('%Y-%m-%d'),
        'desde_display':  desde.strftime('%d/%m/%Y'),
        'hasta_display':  hasta.strftime('%d/%m/%Y'),
        'estado_filtro':  estado_filtro,

        'total_servicios_informe':      total_servicios_informe,
        'servicios_terminados_informe': servicios_terminados_informe,
        'servicios_en_proceso_informe': servicios_en_proceso_informe,
        'ingresos_servicios_informe':   round(ingresos_servicios_informe, 0),

        'total_pagos':            total_pagos,
        'monto_pagos':            monto_pagos,
        'total_nomina_registros': total_nomina_registros,
        'monto_nomina':           monto_nomina,

        'labels_servicios': json.dumps(labels_servicios),
        'datos_servicios':  json.dumps(servicios_por_dia),
        'labels_pagos':     json.dumps(labels_pagos),
        'datos_pagos':      json.dumps(pagos_por_dia),
        'labels_nomina':    json.dumps(labels_nomina),
        'datos_nomina':     json.dumps(nomina_por_dia),

        'servicios_tabla': servicios_tabla,
        'pagos_tabla':     pagos_tabla,
        'nomina_tabla':    nomina_tabla,

        'es_gerente': es_gerente(request.user),
    }

    return render(request, 'estadisticas.html', context)


# ─────────────────────────────────────────────────────────────
# APIs JSON para contadores en tiempo real
# ─────────────────────────────────────────────────────────────
@login_required(login_url=reverse_lazy('login:login'))
def api_contador_insumos(request):
    return JsonResponse({'total_insumos': Insumos.objects.filter(estado=True).count()})

@login_required(login_url=reverse_lazy('login:login'))
def api_contador_vehiculos(request):
    return JsonResponse({'total_vehiculos': Vehiculo.objects.filter(estado=True).count()})

@login_required(login_url=reverse_lazy('login:login'))
def api_contador_clientes(request):
    return JsonResponse({'total_clientes': Cliente.objects.filter(estado=True).count()})

@login_required(login_url=reverse_lazy('login:login'))
def api_contador_gastos(request):
    return JsonResponse({'total_gastos': Gastos.objects.filter(estado=True).count()})


# ─────────────────────────────────────────────────────────────
# Exportar Excel
# ─────────────────────────────────────────────────────────────
def exportar_informe_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    desde, hasta, _ = get_rango_fechas(request)
    estado_filtro   = request.GET.get('estado', 'todos')
    modulo          = request.GET.get('modulo', 'servicios')

    wb = openpyxl.Workbook()
    ws = wb.active
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    if modulo == 'servicios':
        ws.title = 'Servicios'
        qs = DetalleServicio.objects.filter(
            estado=True,
            fecha_creacion__date__gte=desde,
            fecha_creacion__date__lte=hasta,
        )
        if estado_filtro != 'todos':
            qs = qs.filter(proceso=estado_filtro)
        headers = ['ID', 'Vehículo', 'Cliente', 'Empleado', 'Fecha', 'Estado', 'Total']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for s in qs.select_related('id_vehiculo', 'cliente', 'empleado').order_by('-fecha_creacion'):
            ws.append([
                s.id,
                str(s.id_vehiculo.placa),
                str(s.cliente.nombre),
                str(s.empleado.user.get_full_name() if s.empleado else '—'),
                s.fecha_creacion.strftime('%d/%m/%Y'),
                s.get_proceso_display(),
                float(s.subtotal),
            ])

    elif modulo == 'pagos':
        ws.title = 'Pagos'
        qs = Pagos.objects.filter(estado=True, fecha__gte=desde, fecha__lte=hasta)
        headers = ['ID', 'Proveedor', 'Fecha', 'Tipo Pago', 'Monto Total']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for p in qs.select_related('proveedor').order_by('-fecha'):
            ws.append([
                p.id_pago, p.proveedor.nombre,
                p.fecha.strftime('%d/%m/%Y'),
                p.tipo_pago, float(p.monto_total),
            ])

    elif modulo == 'nomina':
        ws.title = 'Nómina'
        qs = Nomina.objects.filter(fecha_pago__gte=desde, fecha_pago__lte=hasta)
        headers = ['ID', 'Empleado', 'Fecha Pago', 'Monto']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for n in qs.select_related('empleado__user').order_by('-fecha_pago'):
            ws.append([
                n.id,
                n.empleado.user.get_full_name() or n.empleado.user.username,
                n.fecha_pago.strftime('%d/%m/%Y'),
                float(n.monto),
            ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="informe_{modulo}_{desde}_{hasta}.xlsx"'
    )
    wb.save(response)
    return response